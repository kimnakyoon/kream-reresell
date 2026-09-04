"""판매 내역을 정산 시트 모양의 엑셀(바탕화면\\KREAM 내역 YYYY-MM.xlsx) 로 쓴다.

열 구성은 사용자가 쓰던 '리리셀 정산 시트' 와 같다:
  # | 매입 주문번호 | 제품코드 | 제품이름 (한/영) | 사이즈 | 매입 일시 | 매입가 | 보관판매 주문번호 | 판매 일시 | 정산가 | 수익 | 메모
수익 = 정산가 − 매입가, 메모 = 수익 × 0.3 (원본 시트의 수식 그대로).
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import DATA_DIR
from .history import HistoryResult, SaleRecord

log = logging.getLogger(__name__)

# 원본 시트의 12개 열 + 맨 오른쪽 '비고' (정산 상태 / 짝 맞춘 방법 - 원본 열은 건드리지 않는다)
COLUMNS = [
    ("#", 6), ("매입 주문번호", 18), ("제품코드", 16), ("제품이름 (한/영)", 46), ("사이즈", 10),
    ("매입 일시", 12), ("매입가", 13), ("보관판매 주문번호", 18), ("판매 일시", 12), ("정산가", 13),
    ("수익", 13), ("메모", 14), ("비고", 30),
]
_COL = {title: i for i, (title, _) in enumerate(COLUMNS, start=1)}
MONEY_FORMAT = '#,###"원"'
DATE_FORMAT = 'm"월" d"일"'
SHARE_RATE = 0.3     # 메모 열: 수익 × 0.3 (원본 시트 수식)

FONT = Font(name="Arial", size=10)
HEAD_FONT = Font(name="Arial", size=11, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
UNMATCHED_FILL = PatternFill("solid", fgColor="FFF2CC")


def desktop_dir() -> Path:
    desktop = Path.home() / "Desktop"
    if not desktop.exists():  # OneDrive 등으로 바탕화면이 옮겨진 경우
        one = Path.home() / "OneDrive" / "Desktop"
        desktop = one if one.exists() else DATA_DIR
    return desktop


def default_path(year: int, month: int) -> Path:
    return desktop_dir() / f"KREAM 내역 {year}-{month:02d}.xlsx"


def _code(style_code: str):
    """숫자로만 된 제품코드(레고 42154 등)는 숫자로 넣는다 (원본 시트와 같게)."""
    if not style_code:
        return None
    return int(style_code) if style_code.isdigit() else style_code


def _note(s: SaleRecord) -> str:
    parts = []
    if s.paid:
        parts.append(f"정산완료 {s.paid_at:%m/%d}" if s.paid_at else "정산완료")
    else:
        parts.append(f"{s.status or '판매완료'} - 정산 전 (정산 예정 금액)")
    if s.purchase is None:
        parts.append("구매 내역에서 매입 건을 찾지 못함")
    elif s.match_how == "상품·사이즈":
        parts.append("상품명·사이즈로 짝 맞춤 (보관번호 없음)")
    elif s.purchase.price is None and s.purchase.total_price is None:
        parts.append("매입가를 읽지 못함")
    elif s.purchase.total_price is None:
        parts.append("결제금액을 못 읽어 매입가에 즉시 구매가(수수료 제외)를 넣음")
    return " · ".join(parts)


def _purchase_price(p) -> int | None:
    """매입가는 항상 수수료를 포함한 실제 결제금액 (즉시 구매가 + 수수료·검수비·배송비). 사용자 확인 2026-09-04.

    결제금액을 못 읽은 경우에만 즉시 구매가로 대신한다 (비고에 표시).
    """
    if p is None:
        return None
    return p.total_price if p.total_price is not None else p.price


def _row_values(n: int, s: SaleRecord) -> list:
    p = s.purchase
    return [
        n,
        p.oid if p and p.oid else (f"입찰 #{p.bid_id}" if p else None),
        _code(s.style_code),
        s.name,
        s.option or None,
        p.ordered_at.replace(hour=0, minute=0, second=0, microsecond=0) if p and p.ordered_at else None,
        _purchase_price(p),
        s.oid or None,
        s.sold_at.replace(hour=0, minute=0, second=0, microsecond=0) if s.sold_at else None,
        s.payout,
        None, None,          # 수익 / 메모 는 수식
        _note(s),
    ]


def write_history(result: HistoryResult, path: Path | None = None) -> Path:
    """엑셀로 저장하고 경로를 돌려준다. 같은 이름의 파일이 열려 있어 못 쓰면 시각을 붙인 이름으로 저장한다."""
    path = path or default_path(result.year, result.month)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{result.month}월"

    for col, (title, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 15.75
    ws.freeze_panes = "A2"

    for i, s in enumerate(result.sales, start=1):
        row = i + 1
        values = _row_values(i, s)
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v)
        g, j, k = (get_column_letter(_COL[t]) for t in ("매입가", "정산가", "수익"))
        if _purchase_price(s.purchase) is not None:
            ws.cell(row=row, column=_COL["수익"], value=f"={j}{row}-{g}{row}")
            ws.cell(row=row, column=_COL["메모"], value=f"={k}{row}*{SHARE_RATE:g}")
        for col in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=row, column=col)
            c.font = FONT
            c.border = BORDER
            if s.purchase is None:
                c.fill = UNMATCHED_FILL
        ws.cell(row=row, column=_COL["비고"]).font = Font(name="Arial", size=9, color="666666")
        for t in ("매입가", "정산가", "수익", "메모"):
            ws.cell(row=row, column=_COL[t]).number_format = MONEY_FORMAT
        for t in ("매입 일시", "판매 일시"):
            ws.cell(row=row, column=_COL[t]).number_format = DATE_FORMAT
        ws.cell(row=row, column=_COL["#"]).alignment = Alignment(horizontal="center")

    # 합계 줄 (원본 시트에는 없지만 한눈에 보기 위해 한 줄 띄우고 둔다)
    if result.sales:
        last = len(result.sales) + 1
        total_row = last + 2
        ws.cell(row=total_row, column=_COL["제품이름 (한/영)"], value=f"합계 ({len(result.sales)}건)").font = HEAD_FONT
        for t in ("매입가", "정산가", "수익", "메모"):
            col = get_column_letter(_COL[t])
            c = ws.cell(row=total_row, column=_COL[t], value=f"=SUM({col}2:{col}{last})")
            c.number_format = MONEY_FORMAT
            c.font = HEAD_FONT
        note = (f"{result.year}년 {result.month}월 거래일시(보관 판매 > 보관 상세) 기준 · "
                f"매입가 = 구매 상세의 결제금액(즉시 구매가 + 수수료) · 수익 = 정산가 − 매입가 · 메모 = 수익 × {SHARE_RATE:g} · "
                f"노란 줄 = 구매 내역에서 짝을 못 찾은 판매 · 만든 시각 {datetime.now():%Y-%m-%d %H:%M}")
        ws.cell(row=total_row + 1, column=_COL["제품이름 (한/영)"], value=note).font = Font(name="Arial", size=9, color="888888")

    try:
        wb.save(path)
    except PermissionError:
        alt = path.with_name(f"{path.stem} {datetime.now():%H%M%S}{path.suffix}")
        log.warning("%s 을 쓸 수 없어(열려 있음?) %s 로 저장", path.name, alt.name)
        wb.save(alt)
        path = alt
    log.info("내역 엑셀 저장: %s", path)
    return path
