"""실행 결과를 엑셀 보고서(바탕화면\\KREAM 결과\\KREAM 입찰결과 날짜.xlsx) 로 정리한다."""

from __future__ import annotations

import logging
import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import DATA_DIR

log = logging.getLogger(__name__)

# 보고서는 바탕화면의 "KREAM 결과" 폴더에 쌓는다 (.env REPORT_DIR 로 바꿀 수 있다)
def _default_report_dir() -> Path:
    custom = os.environ.get("REPORT_DIR", "").strip()
    if custom:
        return Path(custom)
    desktop = Path.home() / "Desktop"
    if not desktop.exists():  # OneDrive 등으로 바탕화면이 옮겨진 경우
        one = Path.home() / "OneDrive" / "Desktop"
        desktop = one if one.exists() else DATA_DIR
    return desktop / "KREAM 결과"


REPORT_DIR = _default_report_dir()


@dataclass
class ProductResult:
    rank: int
    product_id: int
    name: str
    url: str
    category: str = ""          # 랭킹 상품군 (가방, 신발 ...)
    status: str = ""            # 입찰완료 / 입찰대상(dry-run) / 건너뜀 / 중단 / 오류 / 확인필요
    detail: str = ""            # 사유
    fast_sales: int | None = None      # 기간 내 빠른배송 체결 수
    total_sales: int | None = None     # 기간 내 전체 체결 수
    price_a: int | None = None         # 빠른배송 가격 (예상 판매가)
    price_b: int | None = None         # 즉시 판매가
    margin_min: float | None = None    # 이 상품(A 금액 구간)에 적용된 최소 마진율 (0.10 = 10%)
    bid_price: int | None = None       # 입찰가 (= B)
    bid_days: int | None = None
    time: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    @property
    def margin(self) -> int | None:
        if self.price_a is None or self.price_b is None:
            return None
        return self.price_a - self.price_b

    @property
    def margin_rate(self) -> float | None:
        if self.margin is None or not self.price_a:
            return None
        return self.margin / self.price_a


COLUMNS = [
    ("상품군", 10), ("순위", 6), ("상품명", 46), ("상품ID", 10), ("판정", 12), ("사유 / 결과", 46),
    ("30일 빠른배송", 13), ("30일 전체", 10), ("A 빠른배송가", 14), ("B 즉시판매가", 14),
    ("A−B", 11), ("마진율", 9), ("기준마진", 9), ("입찰가", 12), ("입찰기한", 9), ("처리시각", 20), ("링크", 40),
]
_COL = {title: i for i, (title, _) in enumerate(COLUMNS, start=1)}   # 제목 -> 열 번호
_MONEY_COLS = [_COL[t] for t in ("A 빠른배송가", "B 즉시판매가", "A−B", "입찰가")]

STATUS_FILL = {
    "입찰완료": "C6EFCE",
    "입찰대상": "FFEB9C",
    "입찰취소": "C6EFCE",
    "취소대상": "FFEB9C",
    "확인필요": "F8CBAD",
    "오류": "FFC7CE",
    "중단": "FFC7CE",
}

BID_LEGEND = ("판정: 입찰완료 = 실제 입찰됨 / 입찰대상 = dry-run에서 조건 충족 / "
              "건너뜀 = 이미 입찰 중, 조건 미달, 또는 입찰을 시도했지만 넣지 못함 / 확인필요 = 마이페이지에서 입찰 여부 확인")
CANCEL_LEGEND = ("판정: 입찰취소 = 조건 미달이라 입찰을 지움 / 취소대상 = dry-run에서 조건 미달 / 입찰유지 = 조건 충족 / "
                 "확인필요 = 판단 불가 또는 지웠는지 불확실 - 마이페이지에서 확인")


def write_report(results: list[ProductResult], settings_line: str, mode: str,
                 path: Path | None = None, kind: str = "입찰") -> Path:
    """kind 는 파일 이름과 판정 설명에 쓴다: '입찰' 또는 '입찰취소'."""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = path or REPORT_DIR / f"KREAM {kind}결과 {datetime.now():%Y-%m-%d %H%M}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "결과"

    ws["A1"] = f"KREAM 리리셀 {kind} 결과 - {mode}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = settings_line
    ws["A3"] = f"실행 시각: {datetime.now():%Y-%m-%d %H:%M}  |  상품 {len(results)}개"
    ws["A4"] = CANCEL_LEGEND if kind == "입찰취소" else BID_LEGEND
    ws["A4"].font = Font(color="666666", size=9)

    header_row = 6
    bold = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="222222")
    for col, (title, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=header_row, column=col, value=title)
        c.font = bold
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for i, r in enumerate(results, start=header_row + 1):
        values = [
            r.category, r.rank, r.name, r.product_id, r.status, r.detail,
            r.fast_sales, r.total_sales, r.price_a, r.price_b,
            r.margin, r.margin_rate, r.margin_min, r.bid_price,
            f"{r.bid_days}일" if r.bid_days else None, r.time, r.url,
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=i, column=col, value=v)
        for col in _MONEY_COLS:
            ws.cell(row=i, column=col).number_format = "#,##0"
        ws.cell(row=i, column=_COL["마진율"]).number_format = "0.0%"
        ws.cell(row=i, column=_COL["기준마진"]).number_format = "0.0%"
        link = ws.cell(row=i, column=_COL["링크"])
        link.hyperlink = r.url
        link.font = Font(color="0563C1", underline="single")
        fill = STATUS_FILL.get(r.status)
        if fill:
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=fill)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + max(len(results), 1)}"

    # 요약 시트: 판정별 합계 + 상품군별 판정 수
    ss = wb.create_sheet("요약")
    counts: dict[str, int] = {}
    for r in results:
        counts[r.status] = counts.get(r.status, 0) + 1
    ss["A1"] = "판정"
    ss["B1"] = "상품 수"
    ss["A1"].font = ss["B1"].font = Font(bold=True)
    row = 2
    for k, v in sorted(counts.items()):
        ss.cell(row=row, column=1, value=k)
        ss.cell(row=row, column=2, value=v)
        row += 1

    categories: list[str] = []
    for r in results:
        if r.category and r.category not in categories:
            categories.append(r.category)
    if len(categories) > 1:
        row += 1
        statuses = sorted(counts)
        ss.cell(row=row, column=1, value="상품군").font = Font(bold=True)
        for j, st in enumerate(statuses, start=2):
            ss.cell(row=row, column=j, value=st).font = Font(bold=True)
        for cat in categories:
            row += 1
            ss.cell(row=row, column=1, value=cat)
            for j, st in enumerate(statuses, start=2):
                ss.cell(row=row, column=j, value=sum(1 for r in results if r.category == cat and r.status == st))
    ss.column_dimensions["A"].width = 14

    wb.save(path)
    log.info("엑셀 보고서 저장: %s", path)
    return path


def open_file(path: Path) -> None:
    """윈도우 기본 프로그램(엑셀)으로 연다."""
    try:
        os.startfile(str(path))  # type: ignore[attr-defined]
    except Exception as e:  # noqa: BLE001
        log.warning("파일을 열지 못했습니다(%s): %s", e, path)
